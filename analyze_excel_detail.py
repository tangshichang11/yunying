#!/usr/bin/env python3
"""Deep Excel Analysis - Focus on Summary and Calculation Sheets"""

import openpyxl
from openpyxl.utils import get_column_letter
import json

def analyze_sheet_detail(filepath, sheet_name, max_rows=100):
    """Analyze a specific sheet in detail."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb[sheet_name]

    result = {
        "sheet": sheet_name,
        "dimensions": f"{ws.max_row} rows x {ws.max_column} cols",
        "headers": [],
        "rows": []
    }

    # Get headers (first row or row with column headers)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    result["headers"] = [str(h)[:30] if h else "" for h in headers[:20]]

    # Get first several rows of data
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        if any(cell is not None for cell in row):
            result["rows"].append({
                "row": row_idx + 1,
                "values": [str(v)[:40] if v else "" for v in row[:20]]
            })

    wb.close()
    return result

def main():
    # Analyze key sheets from 龙口悦致
    print("="*80)
    print("DEEP ANALYSIS: 龙口悦致.xlsx")
    print("="*80)

    key_sheets = [
        "汇总表",
        "收入及其他项目",
        "阿米巴核算表（日）",
        "阿米巴核算表（月）",
        "单间成本表"
    ]

    for sheet in key_sheets:
        try:
            result = analyze_sheet_detail("data/source/龙口悦致.xlsx", sheet)
            print(f"\n{'='*60}")
            print(f"Sheet: {result['sheet']}")
            print(f"Dimensions: {result['dimensions']}")
            print(f"{'='*60}")
            print(f"Headers (first 20): {result['headers']}")
            print(f"\nData Rows:")
            for row in result['rows'][:15]:
                print(f"  Row {row['row']}: {row['values']}")
        except Exception as e:
            print(f"Error analyzing {sheet}: {e}")

    # Analyze 运营部日核算表
    print("\n" + "="*80)
    print("DEEP ANALYSIS: 运营部日核算表-2026年8月.xlsx")
    print("="*80)

    try:
        result = analyze_sheet_detail("data/source/运营部日核算表-2026年8月.xlsx", "8月日核算表")
        print(f"\nSheet: {result['sheet']}")
        print(f"Dimensions: {result['dimensions']}")
        print(f"Headers: {result['headers']}")
        print(f"\nData Rows:")
        for row in result['rows'][:20]:
            print(f"  Row {row['row']}: {row['values']}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    main()
