#!/usr/bin/env python3
"""Excel Analysis Script for Huiyou Hotel Accounting"""

import openpyxl
from openpyxl.utils import get_column_letter
import json
from pathlib import Path

def analyze_workbook(filepath):
    """Analyze a single workbook and return detailed information."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    result = {
        "filepath": str(filepath),
        "sheets": []
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "cells_with_formulas": [],
            "cells_with_values": [],
            "first_rows": [],
            "has_formulas": False,
            "used_ranges": []
        }

        # Collect cells with formulas
        formula_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        if formula_count < 20:  # Limit to first 20 formulas
                            sheet_info["cells_with_formulas"].append({
                                "cell": cell.coordinate,
                                "formula": cell.value[:200]
                            })
                        formula_count += 1
                        sheet_info["has_formulas"] = True

        # First few rows preview
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
            if any(cell is not None for cell in row):
                sheet_info["first_rows"].append({
                    "row": row_idx + 1,
                    "values": [str(v)[:50] if v else None for v in row[:15]]
                })

        sheet_info["total_formulas"] = formula_count
        result["sheets"].append(sheet_info)

    wb.close()
    return result

def main():
    base_path = Path("data/source")

    files = [
        "龙口悦致.xlsx",
        "运营部日核算表-2026年8月.xlsx"
    ]

    for filename in files:
        filepath = base_path / filename
        print(f"\n{'='*80}")
        print(f"FILE: {filename}")
        print(f"{'='*80}")

        try:
            result = analyze_workbook(filepath)
            print(f"Total Sheets: {len(result['sheets'])}")
            print(f"\nSheet Names:")
            for i, sheet in enumerate(result['sheets']):
                print(f"  {i+1}. {sheet['name']} (rows: {sheet['max_row']}, cols: {sheet['max_column']})")

            for sheet in result['sheets']:
                print(f"\n{'-'*60}")
                print(f"Sheet: {sheet['name']}")
                print(f"{'-'*60}")
                print(f"Dimensions: {sheet['max_row']} rows x {sheet['max_column']} columns")
                print(f"Has Formulas: {sheet['has_formulas']} (Total: {sheet['total_formulas']})")

                if sheet['cells_with_formulas']:
                    print(f"\nSample Formulas (first {len(sheet['cells_with_formulas'])}):")
                    for f in sheet['cells_with_formulas'][:5]:
                        print(f"  {f['cell']}: {f['formula']}")

                if sheet['first_rows']:
                    print(f"\nFirst Rows Preview:")
                    for row in sheet['first_rows'][:5]:
                        print(f"  Row {row['row']}: {row['values']}")

        except Exception as e:
            print(f"Error: {e}")

if __name__ == "__main__":
    main()
